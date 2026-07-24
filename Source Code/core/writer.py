from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter

from core.validator import ColumnValidation
from utils.normalizer import clean_text


YES_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
NO_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")


class WorkbookWriter:
    def __init__(self, add_comments: bool = True):
        self.add_comments = add_comments

    def write(self, wb: Workbook, validations: list[ColumnValidation], output_path: str | Path) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        by_sheet: dict[str, list[ColumnValidation]] = {}
        for validation in validations:
            by_sheet.setdefault(validation.filename_column.sheet_name, []).append(validation)

        for sheet_name, sheet_validations in by_sheet.items():
            ws = wb[sheet_name]
            for validation in sorted(
                sheet_validations,
                key=lambda item: item.filename_column.column,
                reverse=True,
            ):
                self._write_column(ws, validation)

        wb.save(output)
        return output

    def _write_column(self, ws, validation: ColumnValidation) -> None:
        source_col = validation.filename_column.column
        # target_col = source_col + 1
        status_col = source_col + 1
        remarks_col = source_col + 2
        header_row = validation.filename_column.header_row
        header = self._status_header(validation.filename_column.header)

        # if clean_text(ws.cell(header_row, target_col).value) != header:
        #     ws.insert_cols(target_col)
        #     self._copy_column_format(ws, source_col, target_col)
        ws.insert_cols(status_col, amount=2)

        self._copy_column_format(ws, source_col, status_col)
        self._copy_column_format(ws, source_col, remarks_col)

        # header_cell = ws.cell(row=header_row, column=target_col)
        # self._copy_cell_format(ws.cell(row=header_row, column=source_col), header_cell)
        # header_cell.value = header
        status_header = ws.cell(row=header_row, column=status_col)
        remarks_header = ws.cell(row=header_row, column=remarks_col)

        self._copy_cell_format(ws.cell(header_row, source_col), status_header)
        self._copy_cell_format(ws.cell(header_row, source_col), remarks_header)

        status_header.value = header
        remarks_header.value = "Remarks on Files Found"

        for result in validation.results:
            # status_cell = ws.cell(row=result.row, column=target_col)
            # self._copy_cell_format(ws.cell(row=result.row, column=source_col), status_cell)
            # status_cell.value = result.status
            # status_cell.fill = copy(YES_FILL if result.status == "Yes" else NO_FILL)
            # if self.add_comments:
            #     status_cell.comment = self._comment_for(result)
            
            status_cell = ws.cell(row=result.row, column=status_col)
            remarks_cell = ws.cell(row=result.row, column=remarks_col)

            self._copy_cell_format(
                ws.cell(result.row, source_col),
                status_cell,
            )

            self._copy_cell_format(
                ws.cell(result.row, source_col),
                remarks_cell,
            )

            status_cell.value = result.status

            if result.status == "Yes":
                status_cell.fill = copy(YES_FILL)

            elif result.status == "No":
                status_cell.fill = copy(NO_FILL)

            else:
                # Blank marker (/////, NA, etc.)
                status_cell.fill = copy(ws.cell(result.row, source_col).fill)
            if result.status == "":
                remarks_cell.value = ""

            elif result.status == "Yes":
                remarks_cell.value = "Found"

            else:
                remarks = []

                if result.missing_references:
                    remarks.append("Missing:")
                    remarks.extend(result.missing_references)

                # if result.matched_paths:
                #     remarks.append("")
                #     remarks.append("Matched:")
                #     remarks.extend(result.matched_paths)

                remarks_cell.value = "\n".join(remarks)
            # if result.status == "Yes":

            #     remarks_cell.value = "Found"

            # else:

            #     remarks = []

            #     if result.missing_references:

            #         remarks.append("Missing:")

            #         remarks.extend(result.missing_references)

                # if result.matched_paths:

                #     remarks.append("")

                #     remarks.append("Matched:")

                #     remarks.extend(result.matched_paths)

                remarks_cell.value = "\n".join(remarks)

            remarks_cell.alignment = copy(status_cell.alignment)
            remarks_cell.alignment = remarks_cell.alignment.copy(wrapText=True)

            if self.add_comments:
                status_cell.comment = self._comment_for(result)
            
            remarks_letter = get_column_letter(remarks_col)

            ws.column_dimensions[remarks_letter].width = 60

    @staticmethod
    def _status_header(source_header: str) -> str:
        return f"Found? - {source_header or 'Referenced File'}"

    @staticmethod
    def _copy_column_format(ws, source_col: int, target_col: int) -> None:
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)
        source_dim = ws.column_dimensions[source_letter]
        target_dim = ws.column_dimensions[target_letter]
        target_dim.width = source_dim.width or 14
        target_dim.hidden = False
        target_dim.outlineLevel = source_dim.outlineLevel

    @staticmethod
    def _copy_cell_format(source, target) -> None:
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)

    @staticmethod
    def _comment_for(result) -> Comment | None:
        lines: list[str] = []
        if result.missing_references:
            lines.append("Missing:")
            lines.extend(f"- {item}" for item in result.missing_references[:20])
        # if result.matched_paths:
        #     lines.append("Matched:")
        #     lines.extend(f"- {item}" for item in result.matched_paths[:20])
        if not lines:
            return None
        return Comment("\n".join(lines), "Compliance Validator")
