from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.normalizer import clean_text, normalize_token


@dataclass(frozen=True)
class FilenameColumn:
    sheet_name: str
    header_row: int
    column: int
    header: str


class WorkbookInspector:
    def __init__(self, workbook_path: str | Path, max_header_scan_rows: int = 80):
        self.workbook_path = Path(workbook_path)
        self.max_header_scan_rows = max_header_scan_rows

    def load(self) -> Workbook:
        return load_workbook(self.workbook_path, data_only=False)

    def find_filename_columns(self, wb: Workbook) -> list[FilenameColumn]:
        columns: list[FilenameColumn] = []
        for ws in wb.worksheets:
            columns.extend(self._find_sheet_columns(ws))
        return columns

    def _find_sheet_columns(self, ws: Worksheet) -> list[FilenameColumn]:
        found_by_col: dict[int, FilenameColumn] = {}
        max_row = min(ws.max_row or 1, self.max_header_scan_rows)

        for row in ws.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                header = self._merged_aware_value(ws, cell)
                score = self._header_score(header)
                if score < 5:
                    continue

                current = found_by_col.get(cell.column)
                candidate = FilenameColumn(
                    sheet_name=ws.title,
                    header_row=cell.row,
                    column=cell.column,
                    header=clean_text(header),
                )
                if current is None or score > self._header_score(current.header):
                    found_by_col[cell.column] = candidate

        return sorted(found_by_col.values(), key=lambda item: (item.header_row, item.column))

    @staticmethod
    def _merged_aware_value(ws: Worksheet, cell: Cell) -> object:
        if cell.value is not None:
            return cell.value
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return None

    @staticmethod
    def _header_score(value: object) -> int:
        text = normalize_token(clean_text(value))
        if not text:
            return 0

        compact = re.sub(r"[^a-z0-9]+", " ", text).strip()
        if not compact:
            return 0
        if len(compact) > 120 and any(word in compact for word in ("provide", "ensure", "format", "checks")):
            return 0

        reject_words = {
            "status",
            "availability",
            "available",
            "yes no",
            "yes/no",
            "remarks",
            "remark",
            "comment",
            "comments",
        }
        if any(word in compact for word in reject_words):
            return 0

        score = 0
        strong_phrases = (
            "file name",
            "filename",
            "ref document",
            "reference document",
            "document name",
            "doc name",
            "folder name",
            "file folder name",
            "name of file",
            "name of document",
            "ref document name",
            "reference document name",
        )
        if any(phrase in compact for phrase in strong_phrases):
            score += 5
        if "file" in compact:
            score += 2
        if "document" in compact or "doc" in compact:
            score += 2
        if "folder" in compact:
            score += 2
        if "name" in compact:
            score += 2
        if "ref" in compact or "reference" in compact:
            score += 1

        # Real filenames in data rows often include extensions; headers usually do not.
        if re.search(r"\.[a-z0-9]{2,5}\b", text):
            score -= 3
        return score
